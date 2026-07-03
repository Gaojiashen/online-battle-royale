"""
创建《隐秘爱丁堡》角色卡 Excel 模板
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import os

DATA_DIR = Path(__file__).parent / "data"
EXCEL_PATH = DATA_DIR / "角色卡.xlsx"

# ── 样式 ────────────────────────────────────────────
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUB_HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="2F5496")
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LABEL_FONT = Font(name="Microsoft YaHei", bold=True, size=10)
CONTENT_FONT = Font(name="Microsoft YaHei", size=10)
NOTE_FONT = Font(name="Microsoft YaHei", size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_sub_header(ws, row, col, text, span=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUB_HEADER_FONT
    cell.fill = SUB_HEADER_FILL
    cell.alignment = LEFT_WRAP
    cell.border = THIN_BORDER
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)


def style_label(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = LABEL_FONT
    cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_content(ws, row, col, text=""):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = CONTENT_FONT
    cell.alignment = LEFT_WRAP
    cell.border = THIN_BORDER
    return cell


# ── 创建 ────────────────────────────────────────────
os.makedirs(DATA_DIR, exist_ok=True)

wb = Workbook()

# ═══════════════════════════════════════════════════
# Sheet 1 — 基础信息
# ═══════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "基础信息"
ws1.sheet_properties.tabColor = "2F5496"

# 列宽
col_widths_1 = [18, 22, 18, 22, 10, 22]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 标题
ws1.merge_cells("A1:F1")
title_cell = ws1.cell(row=1, column=1, value="《隐秘爱丁堡》角色卡")
title_cell.font = Font(name="Microsoft YaHei", bold=True, size=16, color="2F5496")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# ── 玩家基础信息 ──
row = 3
style_sub_header(ws1, row, 1, "📋 玩家信息", 6); row += 1

info_fields = [
    ("玩家名称", "", "玩家身份", ""),
    ("当前轮次", "第 1 轮", "游戏状态", "存活 / 受伤 / 死亡"),
    ("所在区域", "", "体力/精力", ""),
    ("背景简述", "", "个人目标", ""),
]

for label1, val1, label2, val2 in info_fields:
    style_label(ws1, row, 1, label1)
    style_content(ws1, row, 2, val1)
    style_label(ws1, row, 3, label2)
    style_content(ws1, row, 4, val2)
    # 备注列
    style_label(ws1, row, 5, "备注")
    style_content(ws1, row, 6, "")
    ws1.row_dimensions[row].height = 28
    row += 1

# ── 性相（六维属性） ──
row += 1
style_sub_header(ws1, row, 1, "✨ 性相（Aspects）", 6); row += 1

# 表头
aspect_headers = ["性相", "等级", "已解锁技能", "性相", "等级", "已解锁技能"]
for col, h in enumerate(aspect_headers, 1):
    c = ws1.cell(row=row, column=col, value=h)
    c.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN_BORDER
row += 1

aspects = [
    ("🕯️ 灯（洞察·知识）", 0, "", "🦋 蛾（变化·隐秘）", 0, ""),
    ("⚔️ 刃（冲突·毁灭）", 0, "", "🔨 铸（锻造·创造）", 0, ""),
    ("❄️ 冬（寂灭·坚韧）", 0, "", "❤️ 心（生命·守护）", 0, ""),
]

for a1, l1, s1, a2, l2, s2 in aspects:
    style_content(ws1, row, 1, a1)
    c = style_content(ws1, row, 2, l1)
    c.alignment = CENTER
    style_content(ws1, row, 3, s1)
    style_content(ws1, row, 4, a2)
    c = style_content(ws1, row, 5, l2)
    c.alignment = CENTER
    style_content(ws1, row, 6, s2)
    ws1.row_dimensions[row].height = 28
    row += 1

# ── 技能解锁进度 ──
row += 1
style_sub_header(ws1, row, 1, "🔓 技能解锁规则", 6); row += 1
skill_note = ws1.cell(row=row, column=1, value="性相达 Lv2 → 解锁技能① | Lv6 → 技能② | Lv10 → 技能③ | Lv15 → 技能④")
skill_note.font = NOTE_FONT
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 2

# ── 持有物品 ──
style_sub_header(ws1, row, 1, "🎒 持有物品", 6); row += 1
item_headers = ["物品名称", "数量", "描述/效果", "来源"]
for col, h in enumerate(item_headers, 1):
    c = ws1.cell(row=row, column=col, value=h)
    c.font = Font(name="Microsoft YaHei", bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN_BORDER
ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
# 覆盖合并后的边框
for col in range(4, 7):
    ws1.cell(row=row, column=col).border = THIN_BORDER
row += 1

for _ in range(8):
    for col in [1, 2, 3]:
        style_content(ws1, row, col, "")
    # 描述栏跨列
    desc_cell = style_content(ws1, row, 4, "")
    ws1.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
    ws1.row_dimensions[row].height = 22
    row += 1

# ═══════════════════════════════════════════════════
# Sheet 2 — 区域与路径
# ═══════════════════════════════════════════════════
ws2 = wb.create_sheet("区域与路径")
ws2.sheet_properties.tabColor = "548235"

col_widths_2 = [20, 20, 40, 40]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:D1")
t = ws2.cell(row=1, column=1, value="📍 区域与路径记录")
t.font = Font(name="Microsoft YaHei", bold=True, size=14, color="548235")
t.alignment = Alignment(horizontal="center")
ws2.row_dimensions[1].height = 30

row = 3
headers2 = ["区域名称", "已发现路径", "区域描述/线索", "探索状态"]
for col, h in enumerate(headers2, 1):
    c = ws2.cell(row=row, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = THIN_BORDER

# 预填地图中的区域
areas = [
    "① 圣吉尔斯大教堂",
    "② 皇家哩大道",
    "③ 大学区",
    "④ 城堡岩",
    "⑤ 修道院街",
    "⑥ 灰衣修士墓园",
    "⑦ 草市场",
    "⑧ 格拉斯市场",
    "⑨ 牛门",
    "⑩ 地下墓穴",
]
for area in areas:
    row += 1
    style_content(ws2, row, 1, area)
    style_content(ws2, row, 2, "")
    style_content(ws2, row, 3, "")
    c = style_content(ws2, row, 4, "🔒 未探索")
    c.font = Font(name="Microsoft YaHei", size=10, color="999999")
    ws2.row_dimensions[row].height = 24

# ═══════════════════════════════════════════════════
# Sheet 3 — NPC 关系
# ═══════════════════════════════════════════════════
ws3 = wb.create_sheet("NPC关系")
ws3.sheet_properties.tabColor = "BF8F00"

col_widths_3 = [18, 10, 35, 35]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:D1")
t = ws3.cell(row=1, column=1, value="👥 NPC 关系记录")
t.font = Font(name="Microsoft YaHei", bold=True, size=14, color="BF8F00")
t.alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 30

row = 3
headers3 = ["NPC 名称", "好感度", "已知信息", "交互记录"]
for col, h in enumerate(headers3, 1):
    c = ws3.cell(row=row, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN_BORDER

# 预填 NPCs
npcs = [
    "伊索贝尔·格雷（Isobel Gray）",
]
for npc in npcs:
    row += 1
    style_content(ws3, row, 1, npc)
    c = style_content(ws3, row, 2, "中立")
    c.alignment = CENTER
    style_content(ws3, row, 3, "")
    style_content(ws3, row, 4, "")
    ws3.row_dimensions[row].height = 28

for _ in range(5):
    row += 1
    for col in range(1, 5):
        style_content(ws3, row, col, "")
    ws3.row_dimensions[row].height = 24

# ═══════════════════════════════════════════════════
# Sheet 4 — 行动日志
# ═══════════════════════════════════════════════════
ws4 = wb.create_sheet("行动日志")
ws4.sheet_properties.tabColor = "843C0C"

col_widths_4 = [10, 16, 10, 40, 30]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:E1")
t = ws4.cell(row=1, column=1, value="📜 行动日志（按轮次记录）")
t.font = Font(name="Microsoft YaHei", bold=True, size=14, color="843C0C")
t.alignment = Alignment(horizontal="center")
ws4.row_dimensions[1].height = 30

row = 3
headers4 = ["轮次", "行动动词", "目标", "方式/详情", "裁定结果"]
for col, h in enumerate(headers4, 1):
    c = ws4.cell(row=row, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN_BORDER

for r in range(1, 21):
    row += 1
    for col in range(1, 6):
        style_content(ws4, row, col, "")
    ws4.row_dimensions[row].height = 28

# ═══════════════════════════════════════════════════
# Sheet 5 — 备忘
# ═══════════════════════════════════════════════════
ws5 = wb.create_sheet("备忘")
ws5.sheet_properties.tabColor = "7030A0"

col_widths_5 = [10, 50]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

ws5.merge_cells("A1:B1")
t = ws5.cell(row=1, column=1, value="📝 玩家备忘")
t.font = Font(name="Microsoft YaHei", bold=True, size=14, color="7030A0")
t.alignment = Alignment(horizontal="center")
ws5.row_dimensions[1].height = 30

row = 3
headers5 = ["#", "内容"]
for col, h in enumerate(headers5, 1):
    c = ws5.cell(row=row, column=col, value=h)
    c.font = HEADER_FONT
    c.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
    c.alignment = CENTER
    c.border = THIN_BORDER

for i in range(1, 16):
    row += 1
    c = style_content(ws5, row, 1, i)
    c.alignment = CENTER
    style_content(ws5, row, 2, "")
    ws5.row_dimensions[row].height = 24


# ── 保存 ────────────────────────────────────────────
wb.save(EXCEL_PATH)
print(f"OK 角色卡已保存: {EXCEL_PATH}")
print(f"  工作表: {', '.join(wb.sheetnames)}")
