"""
创建《隐秘爱丁堡》NPC 数据库 Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

DATA_DIR = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(DATA_DIR, "NPC数据库.xlsx")

# ── 样式 ──
FONT_NAME = "Microsoft YaHei"
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
SUB_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUB_FONT = Font(name=FONT_NAME, bold=True, size=10, color="2F5496")
BODY_FONT = Font(name=FONT_NAME, size=10)
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
LEFT_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)

def hcell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = THIN_BORDER
    return c

def bcell(ws, row, col, text=""):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BODY_FONT; c.alignment = LEFT_WRAP; c.border = THIN_BORDER
    return c

def note_row(ws, row, col, text, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = NOTE_FONT; c.alignment = LEFT_WRAP
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)

# ═══════════════════════════════════════════════════
wb = Workbook()

# =====================================================
# Sheet 1 — NPC总览（一行一个NPC）
# =====================================================
ws1 = wb.active
ws1.title = "NPC总览"
ws1.sheet_properties.tabColor = "2F5496"

cols1 = [
    ("编号", 8),
    ("全名", 18),
    ("年龄", 8),
    ("外观描述", 28),
    ("嗓音/语气", 18),
    ("常在区域", 22),
    ("表层身份", 22),
    ("真实欲望/目标", 30),
    ("性格关键词", 18),
    ("阵营倾向", 10),
    ("独有能力", 22),
    ("弱点", 22),
    ("状态", 10),
    ("设计备注", 22),
]
for i, (name, w) in enumerate(cols1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# 标题
ws1.merge_cells("A1:N1")
c = ws1.cell(row=1, column=1, value="《隐秘爱丁堡》NPC 数据库")
c.font = Font(name=FONT_NAME, bold=True, size=14, color="2F5496")
c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

# 说明行
note_row(ws1, 2, 1, "提示：灰色表头 = 核心必填 | 蓝色背景 = 建议填写 | 白色 = 选填", 14)

# 表头
row = 3
headers = ["编号", "全名", "年龄", "外观描述", "嗓音/语气", "常在区域",
           "表层身份", "真实欲望/目标", "性格关键词", "阵营倾向",
           "独有能力", "弱点", "状态", "设计备注"]
for col, h in enumerate(headers, 1):
    hcell(ws1, row, col, h)
ws1.row_dimensions[row].height = 28

# 空行模板
for _ in range(30):
    row += 1
    for col in range(1, 15):
        bcell(ws1, row, col, "")
    ws1.row_dimensions[row].height = 60

# 冻结首行
ws1.freeze_panes = "A4"

# 数据验证：阵营倾向
dv_alignment = DataValidation(type="list", formula1='"中立,灯,蛾,刃,铸,冬,心"', allow_blank=True)
dv_alignment.error = "请选择有效的阵营倾向"
dv_alignment.errorTitle = "无效输入"
ws1.add_data_validation(dv_alignment)
dv_alignment.add(f"J4:J33")

# 数据验证：状态
dv_status = DataValidation(type="list", formula1='"设计中,已完成,待修改,已废弃"', allow_blank=True)
ws1.add_data_validation(dv_status)
dv_status.add(f"M4:M33")

# =====================================================
# Sheet 2 — 资源详情（一对多：每个NPC可提供多种资源）
# =====================================================
ws2 = wb.create_sheet("资源详情")
ws2.sheet_properties.tabColor = "548235"

cols2 = [("NPC编号", 10), ("NPC全名", 18), ("资源类型", 14),
         ("具体内容", 36), ("获取条件", 36), ("备注", 20)]
for i, (name, w) in enumerate(cols2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.merge_cells("A1:F1")
c = ws2.cell(row=1, column=1, value="🎁 NPC 可提供资源一览")
c.font = Font(name=FONT_NAME, bold=True, size=13, color="548235")
c.alignment = Alignment(horizontal="center")

row = 3
for h, _ in cols2:
    hcell(ws2, row, cols2.index((h, _)) + 1, h)
ws2.row_dimensions[row].height = 26

for _ in range(40):
    row += 1
    for col in range(1, 7):
        bcell(ws2, row, col, "")
    ws2.row_dimensions[row].height = 30

ws2.freeze_panes = "A4"

# 数据验证：资源类型
dv_rtype = DataValidation(type="list",
    formula1='"密传知识,物品,情报,能力/技能,势力通道,地点通道,独特能力"', allow_blank=True)
ws2.add_data_validation(dv_rtype)
dv_rtype.add(f"C4:C43")

# =====================================================
# Sheet 3 — 剧情方向
# =====================================================
ws3 = wb.create_sheet("剧情方向")
ws3.sheet_properties.tabColor = "BF8F00"

cols3 = [("NPC编号", 10), ("NPC全名", 18), ("类型", 10),
         ("剧情线描述", 48), ("触发条件", 32), ("关联NPC/区域", 22)]
for i, (name, w) in enumerate(cols3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:F1")
c = ws3.cell(row=1, column=1, value="🔗 NPC 剧情方向")
c.font = Font(name=FONT_NAME, bold=True, size=13, color="BF8F00")
c.alignment = Alignment(horizontal="center")

row = 3
for h, _ in cols3:
    hcell(ws3, row, cols3.index((h, _)) + 1, h)

dv_plot = DataValidation(type="list", formula1='"主线,支线,隐藏"', allow_blank=True)
ws3.add_data_validation(dv_plot)

for _ in range(30):
    row += 1
    for col in range(1, 7):
        bcell(ws3, row, col, "")
    ws3.row_dimensions[row].height = 40
    dv_plot.add(f"C{row}")

ws3.freeze_panes = "A4"

# =====================================================
# Sheet 4 — NPC关联图
# =====================================================
ws4 = wb.create_sheet("NPC关联")
ws4.sheet_properties.tabColor = "7030A0"

cols4 = [("NPC_A", 18), ("关系类型", 16), ("NPC_B", 18),
         ("关系描述", 40), ("对剧情影响", 30)]
for i, (name, w) in enumerate(cols4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.merge_cells("A1:E1")
c = ws4.cell(row=1, column=1, value="🔗 NPC 间关联关系")
c.font = Font(name=FONT_NAME, bold=True, size=13, color="7030A0")
c.alignment = Alignment(horizontal="center")

row = 3
for h, _ in cols4:
    hcell(ws4, row, cols4.index((h, _)) + 1, h)

dv_rel = DataValidation(type="list",
    formula1='"盟友,敌对,师生,亲属,情报线人,利益关系,恩怨,未知"', allow_blank=True)
ws4.add_data_validation(dv_rel)

for _ in range(20):
    row += 1
    for col in range(1, 6):
        bcell(ws4, row, col, "")
    ws4.row_dimensions[row].height = 30
    dv_rel.add(f"B{row}")

ws4.freeze_panes = "A4"

# ═══════════════════════════════════════════════════
# 填入示范数据：伊索贝尔·格雷
# ═══════════════════════════════════════════════════

# Sheet 1 — 基础信息
iso_row = 4
ws1.cell(row=iso_row, column=1, value="NPC-001").font = BODY_FONT
ws1.cell(row=iso_row, column=1).alignment = CENTER
ws1.cell(row=iso_row, column=1).border = THIN_BORDER
iso_data = [
    "伊索贝尔·格雷（Isobel Gray）",
    "17岁",
    "苍白瘦削，淡褐色长发，双目覆灰白色翳膜（灯飞升后遗症）。行走时习惯侧头以耳代目。",
    "轻声，咬字极为清晰。不笑，很少慌张。",
    "牛门(⑩)、修道院街(⑤)、格拉斯市场(⑧)",
    "爱丁堡大学旁听生",
    "揭露七年前书店飞升事件的真相——想知道飞升是否真的成功、仪式要求是什么、谁在掩盖",
    "安静礼貌、极度好奇、不怕危险、利用但不伤害无辜",
    "中立",
    "过目不忘的记忆：听过的东西几乎不会忘记；能感知灯之仪式残余痕迹",
    "视觉缺失；遭遇灯之力残余时会陷入恍惚",
    "已完成",
    "核心NPC，多条剧情线的交汇点",
]
for col, val in enumerate(iso_data, 2):
    ws1.cell(row=iso_row, column=col, value=val).font = BODY_FONT
    ws1.cell(row=iso_row, column=col).alignment = LEFT_WRAP
    ws1.cell(row=iso_row, column=col).border = THIN_BORDER
ws1.row_dimensions[iso_row].height = 90

# Sheet 2 — 资源
resources = [
    ("NPC-001", "伊索贝尔·格雷", "密传知识", "七年来拼凑的关于「灯」之路径的笔记和摘抄", "信任她的人"),
    ("NPC-001", "伊索贝尔·格雷", "能力/技能", "能「感觉」到灯之仪式的残余痕迹（她自己不完全理解）", "带她到现场"),
    ("NPC-001", "伊索贝尔·格雷", "物品", "烧焦的羊皮纸——上面有一只舔舐星辰的狐狸", "她随身携带，几乎不示人"),
    ("NPC-001", "伊索贝尔·格雷", "情报", "大学图书馆借阅记录中，有几人借阅方向与她重叠——可能还有其他追查者", "情报交换"),
    ("NPC-001", "伊索贝尔·格雷", "地点通道", "可以进入神学院部分资料室（远亲担保）", "合理的理由"),
    ("NPC-001", "伊索贝尔·格雷", "独特能力", "过目不忘：记得书店那天每一个声音细节，包括飞升者最后的话「光即荆棘」", "—"),
]
for i, (npc_id, name, rtype, detail, cond) in enumerate(resources):
    r = 4 + i
    for col, val in enumerate([npc_id, name, rtype, detail, cond, ""], 1):
        ws2.cell(row=r, column=col, value=val).font = BODY_FONT
        ws2.cell(row=r, column=col).alignment = LEFT_WRAP
        ws2.cell(row=r, column=col).border = THIN_BORDER
    ws2.row_dimensions[r].height = 36

# Sheet 3 — 剧情方向
plots = [
    ("NPC-001", "伊索贝尔·格雷", "主线", "书店飞升事件的真相：飞升是否成功？仪式要求是什么？谁在掩盖？", "玩家主动接触伊索贝尔并建立信任"),
    ("NPC-001", "伊索贝尔·格雷", "支线", "大学中还有其他人在查阅灯相关的隐秘文献——他们是敌是友？", "在神学院资料室触发"),
    ("NPC-001", "伊索贝尔·格雷", "支线", "飞升力量在地下留下了一条「光痕」，从牛门延伸到某处", "伊索贝尔感知到残余痕迹后引导玩家"),
    ("NPC-001", "伊索贝尔·格雷", "隐藏", "狐狸符号的真相——那不是书中插图，而是飞升时从虚空烙印到纸上的", "特殊条件触发"),
]
for i, (npc_id, name, ptype, desc, trigger) in enumerate(plots):
    r = 4 + i
    for col, val in enumerate([npc_id, name, ptype, desc, trigger, ""], 1):
        ws3.cell(row=r, column=col, value=val).font = BODY_FONT
        ws3.cell(row=r, column=col).alignment = LEFT_WRAP
        ws3.cell(row=r, column=col).border = THIN_BORDER
    ws3.row_dimensions[r].height = 40

# ── 保存 ──
wb.save(EXCEL_PATH)
print(f"OK NPC数据库已保存: {EXCEL_PATH}")
print(f"  工作表: {', '.join(wb.sheetnames)}")
print(f"  已预填 NPC-001 伊索贝尔·格雷 作为示范")
